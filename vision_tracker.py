from __future__ import annotations

import sqlite3
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


APP_TITLE = "Vision Issue Tracker"
DB_PATH = Path("data") / "vision_issues.db"

LINES = ["1-1", "1-2", "2-1", "2-2"]
INSTRUMENTS = [
    "Pinhole",
    "Pouch Align",
    "Lead",
    "Sealing",
    "Lead Align",
    "Welding(+)",
    "Welding(-)",
]
WORKERS = ["Hojun Kwak", "Kijung Kim", "Jihoon Yun", "Jisub Yun"]
ACTIVE_STATUS_OPTIONS = ["Action Required", "Monitoring"]
STATUS_OPTIONS = ACTIVE_STATUS_OPTIONS + ["Resolved"]
CATEGORY_MAP = {
    "Hardware": ["Camera", "Lighting"],
    "Software": ["Program Crash"],
    "Recipe": ["Overkill", "Underkill"],
    "Camera Grab Fail": [""],
    "Production": [""],
    "Other": [""],
}
CATEGORIES = list(CATEGORY_MAP.keys())


@dataclass(frozen=True)
class IssueInput:
    issue_time: str
    line: str
    instrument: str
    worker: str
    category: str
    subcategory: str
    title: str
    description: str
    status: str = ACTIVE_STATUS_OPTIONS[0]
    resolved_time: str = ""
    resolution_notes: str = ""


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def downtime_duration(issue_time: str, end_time: datetime | None = None) -> str:
    end_time = end_time or datetime.now()
    try:
        start_time = datetime.strptime(issue_time, "%Y-%m-%d %H:%M")
    except ValueError:
        return ""
    minutes = max(0, int((end_time - start_time).total_seconds() // 60))
    hours, remaining_minutes = divmod(minutes, 60)
    return f"{hours:02d}:{remaining_minutes:02d}"


def connect(db_path: Path = DB_PATH) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def initialize_database(db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                issue_time TEXT NOT NULL,
                resolved_time TEXT,
                line TEXT NOT NULL,
                instrument TEXT NOT NULL,
                worker TEXT NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT,
                title TEXT NOT NULL,
                description TEXT,
                status TEXT NOT NULL,
                resolution_notes TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_issues_lookup
            ON issues(status, category, subcategory, line, instrument, issue_time)
            """
        )
        conn.execute("UPDATE issues SET status = 'Action Required' WHERE status = 'Open'")
        conn.execute("UPDATE issues SET status = 'Monitoring' WHERE status = 'In Progress'")
        conn.execute("UPDATE issues SET subcategory = 'Overkill' WHERE subcategory = 'Overkill(False Reject)'")
        conn.execute("UPDATE issues SET subcategory = 'Underkill' WHERE subcategory = 'Underkill(False Accept)'")
        conn.commit()


def validate_issue(issue: IssueInput) -> list[str]:
    errors: list[str] = []
    required = {
        "Issue time": issue.issue_time,
        "Line": issue.line,
        "Instrument": issue.instrument,
        "Worker": issue.worker,
        "Category": issue.category,
        "Title": issue.title,
        "Status": issue.status,
    }
    for label, value in required.items():
        if not value.strip():
            errors.append(f"{label} is required.")
    try:
        datetime.strptime(issue.issue_time, "%Y-%m-%d %H:%M")
    except ValueError:
        errors.append("Issue time must use YYYY-MM-DD HH:MM format.")
    if issue.line not in LINES:
        errors.append("Line is not valid.")
    if issue.instrument not in INSTRUMENTS:
        errors.append("Instrument is not valid.")
    if issue.category not in CATEGORY_MAP:
        errors.append("Category is not valid.")
    if issue.status not in STATUS_OPTIONS:
        errors.append("Status is not valid.")
    allowed_subcategories = CATEGORY_MAP.get(issue.category, [])
    if issue.subcategory and issue.subcategory not in allowed_subcategories:
        errors.append("Subcategory is not valid for the selected category.")
    return errors


def create_issue(issue: IssueInput, db_path: Path = DB_PATH) -> int:
    errors = validate_issue(issue)
    if errors:
        raise ValueError("\n".join(errors))

    with closing(connect(db_path)) as conn:
        cursor = conn.execute(
            """
            INSERT INTO issues (
                created_at, issue_time, resolved_time, line, instrument, worker,
                category, subcategory, title, description, status, resolution_notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now_text(),
                issue.issue_time,
                issue.resolved_time,
                issue.line,
                issue.instrument,
                issue.worker,
                issue.category,
                issue.subcategory,
                issue.title,
                issue.description,
                issue.status,
                issue.resolution_notes,
            ),
        )
        conn.commit()
        return int(cursor.lastrowid)


def update_issue(issue_id: int, issue: IssueInput, db_path: Path = DB_PATH) -> None:
    errors = validate_issue(issue)
    if errors:
        raise ValueError("\n".join(errors))

    with closing(connect(db_path)) as conn:
        conn.execute(
            """
            UPDATE issues
            SET issue_time = ?, resolved_time = ?, line = ?, instrument = ?,
                worker = ?, category = ?, subcategory = ?, title = ?,
                description = ?, status = ?, resolution_notes = ?
            WHERE id = ?
            """,
            (
                issue.issue_time,
                issue.resolved_time,
                issue.line,
                issue.instrument,
                issue.worker,
                issue.category,
                issue.subcategory,
                issue.title,
                issue.description,
                issue.status,
                issue.resolution_notes,
                issue_id,
            ),
        )
        conn.commit()


def resolve_issue(issue_id: int, notes: str = "", db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT issue_time FROM issues WHERE id = ?", (issue_id,)).fetchone()
        duration = downtime_duration(row["issue_time"]) if row else ""
        conn.execute(
            """
            UPDATE issues
            SET status = 'Resolved',
                resolved_time = COALESCE(NULLIF(resolved_time, ''), ?),
                resolution_notes = ?
            WHERE id = ?
            """,
            (duration, notes, issue_id),
        )
        conn.commit()


def set_issue_status(issue_id: int, status: str, db_path: Path = DB_PATH) -> None:
    if status not in STATUS_OPTIONS:
        raise ValueError("Status is not valid.")
    with closing(connect(db_path)) as conn:
        if status == "Resolved":
            row = conn.execute("SELECT issue_time FROM issues WHERE id = ?", (issue_id,)).fetchone()
            duration = downtime_duration(row["issue_time"]) if row else ""
            conn.execute(
                """
                UPDATE issues
                SET status = ?,
                    resolved_time = COALESCE(NULLIF(resolved_time, ''), ?)
                WHERE id = ?
                """,
                (status, duration, issue_id),
            )
        else:
            conn.execute("UPDATE issues SET status = ? WHERE id = ?", (status, issue_id))
        conn.commit()


def delete_issue(issue_id: int, db_path: Path = DB_PATH) -> None:
    with closing(connect(db_path)) as conn:
        conn.execute("DELETE FROM issues WHERE id = ?", (issue_id,))
        conn.commit()


def build_search_query(filters: dict[str, str]) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []

    exact_fields = ["status", "line", "instrument", "category", "subcategory", "worker"]
    for field in exact_fields:
        value = filters.get(field, "").strip()
        if value:
            clauses.append(f"{field} = ?")
            params.append(value)

    date_from = filters.get("date_from", "").strip()
    date_to = filters.get("date_to", "").strip()
    if date_from:
        clauses.append("issue_time >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("issue_time <= ?")
        params.append(date_to)

    keyword = filters.get("keyword", "").strip()
    if keyword:
        clauses.append("(title LIKE ? OR description LIKE ? OR resolution_notes LIKE ?)")
        like_value = f"%{keyword}%"
        params.extend([like_value, like_value, like_value])

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    query = f"""
        SELECT id, issue_time, resolved_time, line, instrument, worker, category,
               subcategory, title, description, status, resolution_notes
        FROM issues
        {where}
        ORDER BY issue_time DESC, id DESC
    """
    return query, params


def active_issues(db_path: Path = DB_PATH) -> list[sqlite3.Row]:
    with closing(connect(db_path)) as conn:
        return list(
            conn.execute(
                """
                SELECT id, issue_time, resolved_time, line, instrument, worker, category,
                       subcategory, title, description, status, resolution_notes
                FROM issues
                WHERE status IN (?, ?)
                ORDER BY issue_time DESC, id DESC
                """,
                tuple(ACTIVE_STATUS_OPTIONS),
            )
        )


def dashboard_counts(db_path: Path = DB_PATH) -> dict[str, int]:
    today = datetime.now().strftime("%Y-%m-%d")
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT status, COUNT(*) AS count
            FROM issues
            GROUP BY status
            """
        ).fetchall()
        counts = {row["status"]: int(row["count"]) for row in rows}
        resolved_today = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM issues
            WHERE status = 'Resolved' AND issue_time >= ? AND issue_time < ?
            """,
            (f"{today} 00:00", f"{today} 23:59"),
        ).fetchone()
        counts["Resolved Today"] = int(resolved_today["count"]) if resolved_today else 0
        counts["Active"] = sum(counts.get(status, 0) for status in ACTIVE_STATUS_OPTIONS)
        return counts


def search_issues(filters: dict[str, str] | None = None, db_path: Path = DB_PATH) -> list[sqlite3.Row]:
    filters = filters or {}
    query, params = build_search_query(filters)
    with closing(connect(db_path)) as conn:
        return list(conn.execute(query, params))


def get_issue(issue_id: int, db_path: Path = DB_PATH) -> sqlite3.Row | None:
    with closing(connect(db_path)) as conn:
        return conn.execute("SELECT * FROM issues WHERE id = ?", (issue_id,)).fetchone()


def export_issues_to_excel(rows: list[sqlite3.Row], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Issue Report"

    headers = [
        "ID",
        "Issue Time",
        "Downtime Duration",
        "Line",
        "Instrument",
        "Logged By",
        "Category",
        "Subcategory",
        "Title",
        "Status",
        "Description",
        "Resolution Notes",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        sheet.append([row[header_key(header)] for header in headers])

    for column_index, header in enumerate(headers, start=1):
        max_length = len(header)
        for cell in sheet[get_column_letter(column_index)]:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 48)

    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def header_key(header: str) -> str:
    return {
        "ID": "id",
        "Issue Time": "issue_time",
        "Downtime Duration": "resolved_time",
        "Line": "line",
        "Instrument": "instrument",
        "Logged By": "worker",
        "Category": "category",
        "Subcategory": "subcategory",
        "Title": "title",
        "Status": "status",
        "Description": "description",
        "Resolution Notes": "resolution_notes",
    }[header]
